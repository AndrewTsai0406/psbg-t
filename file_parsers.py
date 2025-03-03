import re
from tqdm import tqdm
import pandas as pd
from PIL.Image import Image
from pydantic import BaseModel
from IPython.display import display
from typing import List, Tuple, Union, Literal
import openpyxl
from openpyxl.cell.cell import Cell

import camelot
import pdfplumber
from langchain_core.documents import Document
import base64
from io import BytesIO


class PDFElement(BaseModel):
    type: Literal["table", "image", "text"]
    bbox: Tuple[float, float, float, float]  # 邊界框 (left, top, right, bottom)
    page_no: int
    content: Union[str, Image]

    class Config:
        arbitrary_types_allowed = True


class CustomPDFParser:
    def __init__(self):
        pass

    def ignore_key(self, text_line):
        keys_to_ignore = ["chars"]
        text_line = {
            key: value for key, value in text_line.items() if key not in keys_to_ignore
        }
        return text_line

    def is_within_bbox(self, text_bbox, table_bbox):
        """
        判斷文本框是否位於表格框內
        """
        x0, y0, x1, y1 = text_bbox
        tx0, ty0, tx1, ty1 = table_bbox
        return tx0 <= x0 and tx1 >= x1 and ty0 <= y0 and ty1 >= y1

    def is_numbered_list(
        self,
        current_pdf_element: PDFElement,
        last_numbered_list_element: Union[PDFElement, None],
    ) -> bool:
        """
        檢查是否為數字編號模式
        """
        text = current_pdf_element.content

        if text:
            match = re.match(r"^(\d+\.)+\d*", text)
            if match:
                if not last_numbered_list_element:
                    return True  # 第一次出現的編號始終為合法編號

                current_number = match.group(0).strip(".")
                last_number = (
                    re.match(r"^(\d+\.)+\d*", last_numbered_list_element.content)
                    .group(0)
                    .strip(".")
                )

                # 比對編號增長規則
                current_parts = current_number.split(".")
                last_parts = last_number.split(".")
                current_left_bound = current_pdf_element.bbox[0]
                last_left_bound = last_numbered_list_element.bbox[0]
                print("text", text)
                print("current_left_bound", current_left_bound)
                print("last_left_bound", last_left_bound)
                print("-" * 30)
                # 檢查編號是否是合法
                if (
                    len(current_parts) == len(last_parts)
                    and (
                        1 <= int(current_parts[-1]) - int(last_parts[-1]) <= 2
                    )  # 給予1~2容錯
                    and abs(current_left_bound - last_left_bound)
                    <= 2  # 兩者邊界在誤差範圍
                ):
                    # print("Rule1")
                    return True  # 同層遞增 1.2->1.3
                if (
                    len(current_parts) == len(last_parts) + 1
                    and int(current_parts[-1]) == 1
                    and (current_left_bound - last_left_bound) > 2
                ):
                    # print("Rule2")
                    return True  # 進入下一層 1.2->1.2.1
                if (
                    len(current_parts) < len(last_parts)
                    and int(current_parts[len(current_parts) - 1])
                    == int(last_parts[len(current_parts) - 1]) + 1
                    and (last_left_bound - current_left_bound) > 2
                ):
                    # print("Rule3")
                    return True  # 跳層 1.2->2
        return False

    def table_extract_by_camelot(
        self, page_no: int, doc_path: str, page_height: float = 842
    ) -> list[dict]:
        """
        使用camelot擷取table
        Return: [
                    {
                        "bbox": [bbox.l, bbox.t, bbox.r, bbox.b],
                        "df": pd.Dataframe,
                    }
                ]
        """
        tables_with_bbox = []
        tables = camelot.read_pdf(doc_path, pages=str(page_no), flavor="lattice")
        for table in tables:
            l, b, r, t = table._bbox
            top_left_bbox = [l, page_height - t, r, page_height - b]
            tables_with_bbox.append({"bbox": top_left_bbox, "df": table.df})
        return tables_with_bbox

    def contains_japanese(self, text):
        # 日文的Unicode範圍，包括平假名、片假名
        japanese_pattern = re.compile(r"[\u3040-\u30FF\uFF66-\uFF9F]")
        return bool(japanese_pattern.search(text))

    def get_ancestors(
        self, list_items: List[PDFElement], last_list_item: PDFElement
    ) -> List[str]:
        """
        Extracts the ancestors from a list item
        """
        ancestors = []
        number = last_list_item.content

        number_level = re.match(r"^(\d+\.)+\d*", number).group(0).strip(".").split(".")
        for list_item in list_items:
            item_level = (
                re.match(r"^(\d+\.)+\d*", list_item.content)
                .group(0)
                .strip(".")
                .split(".")
            )
            if (
                len(item_level) < len(number_level)
                and number_level[: len(item_level)] == item_level
            ):
                ancestors.append(list_item.content)

        ancestors.append(number)
        return ancestors

    def custom_pdf_parser(self, doc_path: str) -> list[Document]:
        with pdfplumber.open(doc_path) as pdf:
            images_per_page = [len(page.images) >= 1 for page in pdf.pages]
            tables_per_page = [len(page.find_tables()) >= 1 for page in pdf.pages]
            have_jp_per_page = [
                self.contains_japanese(
                    " ".join(
                        [text_line["text"] for text_line in page.extract_text_lines()]
                    )
                )
                for page in pdf.pages
            ]

            is_ppt_pdf = (
                True if all(images_per_page) else False
            )  # PPT轉成PDF會導致每一頁都是images
            # print("Is ppt pdf? ->", is_ppt_pdf)
            is_table_pdf = (
                True if all(tables_per_page) else False
            )  # 有些PPT每一頁都是外框被偵測為PDF
            # print("Is table pdf? ->", is_table_pdf)
            is_jp_pdf = (
                True if all(have_jp_per_page) else False
            )  # 以內文是否為日文判斷是否是SIE
            # print("Is jp pdf? ->", is_jp_pdf)

            elements: List[PDFElement] = []

            for page in tqdm(pdf.pages, desc="Parse pdf"):
                # print(str(page.page_number).center(50, "="))
                # print(page.height)
                page_x0, page_top, page_x1, page_bottom = page.bbox

                page_image = page.to_image(resolution=100)

                table_bboxes = []

                # print("TABLES".center(50, "-"))
                if not is_jp_pdf:
                    tables = page.find_tables(
                        table_settings={
                            "vertical_strategy": "lines",
                            "horizontal_strategy": "lines",
                        }
                    )
                    tables = tables[1:] if is_table_pdf else tables
                    if tables:
                        for table in tables:
                            table_bboxes.append(table.bbox)
                            elements.append(
                                PDFElement(
                                    type="table",
                                    bbox=table.bbox,
                                    page_no=page.page_number,
                                    content=pd.DataFrame(table.extract()).to_markdown(),
                                )
                            )
                            # display(page_image.draw_rect(table.bbox))
                            # display(pd.DataFrame(table.extract()))
                    else:
                        table_bboxes = [(0, 0, 0, 0)]
                else:
                    tables = self.table_extract_by_camelot(
                        page_height=page.height,
                        doc_path=doc_path,
                        page_no=page.page_number,
                    )
                    if tables:
                        for table in tables:
                            table_bboxes.append(table["bbox"])
                            elements.append(
                                PDFElement(
                                    type="table",
                                    bbox=table["bbox"],
                                    page_no=page.page_number,
                                    content=table["df"].to_markdown(),
                                )
                            )
                            # display(page_image.draw_rect(table["bbox"]))
                            # display(table["df"])
                    else:
                        table_bboxes = [(0, 0, 0, 0)]

                # print("IMAGES".center(50, "-"))
                images = page.images[1:] if is_ppt_pdf else page.images
                if images:
                    # print(images[0])
                    for im in images:
                        x0, top, x1, bottom = (
                            im["x0"],
                            im["top"],
                            im["x1"],
                            im["bottom"],
                        )

                        x0 = max(im["x0"], page_x0)
                        top = max(im["top"], page_top)
                        x1 = min(im["x1"], page_x1)
                        bottom = min(im["bottom"], page_bottom)
                        im_pil = (
                            page.crop((x0, top, x1, bottom))
                            .to_image(resolution=72)
                            .original
                        )
                        elements.append(
                            PDFElement(
                                type="image",
                                bbox=(x0, top, x1, bottom),
                                page_no=page.page_number,
                                content=im_pil,
                            )
                        )
                        # display(page_image.draw_rect((x0, top, x1, bottom)))

                # print("TEXTS".center(50, "-"))
                text_lines = page.extract_text_lines()
                for idx, text_line in enumerate(text_lines):
                    is_table_of_contents = re.search(r"\.{5,}", text_line["text"])

                    text_bbox = (
                        text_line["x0"],
                        text_line["top"],
                        text_line["x1"],
                        text_line["bottom"],
                    )

                    if is_table_of_contents or any(
                        self.is_within_bbox(text_bbox, table_bbox)
                        for table_bbox in table_bboxes
                    ):  # 判斷是否是目錄或是在表格內
                        continue

                    elements.append(
                        PDFElement(
                            type="text",
                            bbox=text_bbox,
                            page_no=page.page_number,
                            content=text_line["text"],
                        )
                    )

            elements.sort(key=lambda elem: (elem.page_no, elem.bbox[1], elem.bbox[0]))

            all_numbered_list: List[PDFElement] = []
            documnets: List[Document] = []
            documnet_content: str = ""
            document_metadata = {"image": [], "ancestors": []}

            for element in elements:
                if element.type == "text":
                    if self.is_numbered_list(
                        element,
                        all_numbered_list[-1] if all_numbered_list else None,
                    ):
                        all_numbered_list.append(element)
                        if len(all_numbered_list) == 1 or not documnet_content:
                            continue  # 第一個numbered_list不做後續動作
                        ancestors = self.get_ancestors(
                            all_numbered_list[:-1], all_numbered_list[-2]
                        )
                        document_metadata["ancestors"] = ancestors
                        # documnet_content = "\n".join(ancestors) + documnet_content
                        documnets.append(
                            Document(
                                id=ancestors[-1].strip(),
                                page_content=documnet_content,
                                metadata=document_metadata,
                            )
                        )
                        documnet_content = ""
                        document_metadata = {"image": [], "ancestors": []}
                    else:
                        documnet_content += f"\n{element.content}"
                elif element.type == "table":
                    documnet_content += f"\n{element.content}"
                elif element.type == "image":
                    buffered = BytesIO()
                    element.content.save(buffered, format="PNG")
                    encoded_image = base64.b64encode(buffered.getvalue()).decode(
                        "utf-8"
                    )
                    document_metadata["image"].append(encoded_image)
                    """
                    載回image
                    decoded_image = base64.b64decode(documents[10].metadata['image'][0])
                    image = Image.open(BytesIO(decoded_image))
                    """

            return documnets


def custom_pdf_parser(doc_path: str) -> list[Document]:
    """
    Custom PDF parser to extract text and tables from a PDF file
    """

    def is_same_paragraph(bbox1, bbox2, threshold=10):
        return abs(bbox1[3] - bbox2[3]) < threshold

    with pdfplumber.open(doc_path) as pdf:
        page_contents = []

        for page_num, page in enumerate(pdf.pages):
            text_data = page.extract_words()
            tables = page.extract_tables()

            elements = []
            for word in text_data:
                elements.append(
                    {
                        "type": "text",
                        "content": word["text"],
                        "bbox": (
                            word["x0"],
                            word["top"],
                            word["x1"],
                            word["bottom"],
                        ),
                    }
                )

            for table in tables:
                elements.append(
                    {
                        "type": "table",
                        "content": table,
                        "bbox": (0, 0, 0, 0),
                    }
                )

            elements.sort(key=lambda el: (el["bbox"][1], el["bbox"][0]))
            page_content = []
            current_paragraph = ""
            previous_bbox = None

            for el in elements:
                if el["type"] == "text":
                    current_paragraph += el["content"] + " "
                    current_paragraph += (
                        "\n"
                        if previous_bbox
                        and not is_same_paragraph(previous_bbox, el["bbox"])
                        else ""
                    )
                    previous_bbox = el["bbox"]
                elif el["type"] == "table":
                    if current_paragraph.strip():
                        page_content.append(current_paragraph.strip())
                        current_paragraph = ""
                    page_content.append(pd.DataFrame(el["content"]).to_markdown())

            if current_paragraph.strip():
                page_content.append(current_paragraph.strip())

            page_contents += [x for x in page_content if x]

        page_contents = [Document(page_content) for page_content in page_contents]

    return page_contents


def custom_xlsx_parser(doc_path: str) -> list[Document]:
    """
    xlsx parser, specifically designed for Dell spec documents
    """

    def is_numbered_list(text: str) -> bool:
        if text:
            return bool(re.match(r"^\d+(\.\d+)+", text))
        else:
            return False

    def get_hierarchy_level(text: str) -> int:
        # Counts the hierarchy level based on the dotted numbers (e.g., 1.0 or 1.1)
        numbered_list = re.match(r"^\d+(\.\d+)+", text).group(0)
        numbered_list = numbered_list.split(".")
        level = len(numbered_list) - numbered_list.count("0")
        return level

    def is_column_names(cell: Cell) -> bool:
        # Heuristic: "Parameter Description" or gray fill cells
        if (
            cell.value == "Parameter Description"
            or cell.fill.__getattr__("fgColor").type == "indexed"
        ):
            return True
        else:
            return False

    workbook = openpyxl.load_workbook(doc_path)
    spec_sheet_name = [
        i for i in workbook.sheetnames if re.search(r"\bSpec\b", i, re.IGNORECASE)
    ]
    assert len(spec_sheet_name) == 1
    sheet = workbook[spec_sheet_name[0]]

    current_hierarchy = []
    current_dataframe = pd.DataFrame()
    all_dataframe = []
    column_names = []

    for row in sheet.iter_rows():
        first_cell = row[0]
        if is_numbered_list(first_cell.value):
            # We hit a hierarchy title line
            if current_dataframe.shape[0] > 0:
                all_dataframe.append((current_dataframe, current_hierarchy.copy()))
                current_dataframe = current_dataframe.iloc[0:0]

            level = get_hierarchy_level(first_cell.value)
            current_hierarchy[level - 1 :] = [first_cell.value]

        elif is_column_names(first_cell):
            # We got column names
            column_names = [cell.value for cell in row if cell.value]
            if current_dataframe.shape[0] > 0:
                all_dataframe.append((current_dataframe, current_hierarchy.copy()))
                current_dataframe = current_dataframe.iloc[0:0]
            current_dataframe = pd.DataFrame(columns=column_names)

        else:
            if len(column_names) > 0:
                data = [str(cell.value) if cell.value else cell.value for cell in row][
                    : len(column_names)
                ]
                new_row = pd.DataFrame([data], columns=column_names)
                current_dataframe = pd.concat(
                    [current_dataframe, new_row], ignore_index=True
                )

    if current_dataframe.shape[0] > 0:
        all_dataframe.append((current_dataframe, current_hierarchy.copy()))

    page_contents = []
    document_metadata = {"image": [], "ancestors": []}
    for dataframe, ancestors in all_dataframe:
        # title = "\n".join(ancestors)
        content = dataframe.to_markdown()
        document_metadata["ancestors"] = ancestors
        page_contents.append(
            Document(
                id=ancestors[-1].strip(),
                page_content=content,
                metadata=document_metadata,
            )
        )

    return page_contents
